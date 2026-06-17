import math
from io import BytesIO
from datetime import datetime
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="BOQ Đường dây 110-500kV", page_icon="⚡", layout="wide")

CONDUCTOR_LIB = pd.DataFrame([
    {"Mã dây": "ACSR-185/29", "Tiết diện mm2": 185, "I cho phép A": 430, "R ohm/km": 0.157, "X ohm/km": 0.40, "Khối lượng kg/km": 732, "Ghi chú": "Tham khảo sơ bộ"},
    {"Mã dây": "ACSR-240/32", "Tiết diện mm2": 240, "I cho phép A": 520, "R ohm/km": 0.120, "X ohm/km": 0.39, "Khối lượng kg/km": 922, "Ghi chú": "Tham khảo sơ bộ"},
    {"Mã dây": "ACSR-300/39", "Tiết diện mm2": 300, "I cho phép A": 610, "R ohm/km": 0.097, "X ohm/km": 0.38, "Khối lượng kg/km": 1136, "Ghi chú": "Tham khảo sơ bộ"},
    {"Mã dây": "ACSR-330", "Tiết diện mm2": 330, "I cho phép A": 680, "R ohm/km": 0.088, "X ohm/km": 0.38, "Khối lượng kg/km": 1260, "Ghi chú": "Cần thay bằng catalog/hồ sơ hiện trạng"},
    {"Mã dây": "ACSR-400", "Tiết diện mm2": 400, "I cho phép A": 760, "R ohm/km": 0.073, "X ohm/km": 0.37, "Khối lượng kg/km": 1511, "Ghi chú": "Tham khảo sơ bộ"},
    {"Mã dây": "ACSR-500/64", "Tiết diện mm2": 500, "I cho phép A": 900, "R ohm/km": 0.058, "X ohm/km": 0.36, "Khối lượng kg/km": 1889, "Ghi chú": "Tham khảo sơ bộ"},
    {"Mã dây": "ACSR-630/72", "Tiết diện mm2": 630, "I cho phép A": 1050, "R ohm/km": 0.046, "X ohm/km": 0.35, "Khối lượng kg/km": 2277, "Ghi chú": "Tham khảo sơ bộ"},
])

VOLTAGE_DEFAULTS = {
    "110kV": {"span_ref": 300, "bundle": 1, "jkt": 1.1, "tower_weight": 7.5, "foundation_exc": 28, "foundation_conc": 11, "rebar": 1.2, "anchor": 0.18, "grounding": 0.10},
    "220kV": {"span_ref": 400, "bundle": 2, "jkt": 1.0, "tower_weight": 18.0, "foundation_exc": 65, "foundation_conc": 24, "rebar": 3.2, "anchor": 0.45, "grounding": 0.18},
    "500kV": {"span_ref": 500, "bundle": 4, "jkt": 0.9, "tower_weight": 55.0, "foundation_exc": 180, "foundation_conc": 70, "rebar": 9.0, "anchor": 1.2, "grounding": 0.35},
}

def current_3phase(p_mw, u_kv, cos_phi):
    return p_mw * 1000 / (math.sqrt(3) * u_kv * cos_phi) if u_kv > 0 and cos_phi > 0 else 0

def conductor_from_lib(code):
    df = CONDUCTOR_LIB[CONDUCTOR_LIB["Mã dây"] == code]
    return df.iloc[0] if not df.empty else CONDUCTOR_LIB.iloc[0]

def select_conductor(i_per_wire, s_per_wire):
    df = CONDUCTOR_LIB.copy()
    df["I tính trên 1 dây A"] = i_per_wire
    df["S yêu cầu trên 1 dây mm2"] = s_per_wire
    df["Đạt dòng tải"] = df["I cho phép A"] >= i_per_wire
    df["Đạt tiết diện KT"] = df["Tiết diện mm2"] >= s_per_wire
    df["Đạt sơ bộ"] = df["Đạt dòng tải"] & df["Đạt tiết diện KT"]
    return df

def voltage_drop_percent(i_a, u_kv, length_km, r_ohm_km, x_ohm_km, cos_phi, circuits=1, bundle=1):
    sin_phi = math.sqrt(max(0, 1 - cos_phi**2))
    r_eq = r_ohm_km / max(1, circuits * bundle)
    x_eq = x_ohm_km / max(1, circuits)
    delta_v = math.sqrt(3) * i_a * (r_eq * cos_phi + x_eq * sin_phi) * length_km
    return delta_v / (u_kv * 1000) * 100 if u_kv > 0 else 0

def estimate_towers(length_km, span_m):
    spans = math.ceil(length_km * 1000 / span_m) if span_m > 0 else 0
    return spans, spans + 1

def allocate_by_percent(total_auto, mix_df):
    percent_rows = mix_df[mix_df["Số lượng nhập tay"] <= 0].copy()
    total_percent = percent_rows["Tỷ lệ %"].sum()
    if total_auto <= 0 or total_percent <= 0 or percent_rows.empty:
        return {idx: 0 for idx in percent_rows.index}
    raw = percent_rows["Tỷ lệ %"] / total_percent * total_auto
    floor = raw.apply(math.floor).astype(int)
    remainder = int(total_auto - floor.sum())
    frac = (raw - floor).sort_values(ascending=False)
    alloc = floor.to_dict()
    for idx in frac.index[:remainder]:
        alloc[idx] += 1
    return alloc

def normalize_boq(df):
    out = df.copy()
    out["Khối lượng"] = pd.to_numeric(out["Khối lượng"], errors="coerce").fillna(0)
    out = out[out["Khối lượng"] > 0].reset_index(drop=True)
    out.insert(0, "STT", range(1, len(out) + 1))
    return out

def build_excel(sheets):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])
        wb = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9EAF7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 45)
    return output.getvalue()

def status(ok):
    return "PASS" if ok else "FAIL"

st.sidebar.title("⚡ Input chính")
project_name = st.sidebar.text_input("Tên dự án", "Đường dây đấu nối sơ bộ")
voltage_level = st.sidebar.selectbox("Cấp điện áp", ["110kV", "220kV", "500kV"], index=1)
u_kv = float(voltage_level.replace("kV", ""))
defaults = VOLTAGE_DEFAULTS[voltage_level]
connection_type = st.sidebar.selectbox("Kiểu đấu nối", ["Đấu nối mới", "Transit / cắt xen trên ĐD hiện trạng"], index=1)
p_mw = st.sidebar.number_input("Công suất truyền tải P (MW)", min_value=1.0, value=90.0, step=5.0)
length_km = st.sidebar.number_input("Chiều dài đoạn đấu nối/tuyến mới L (km)", min_value=0.1, value=5.0, step=0.1)
cos_phi = st.sidebar.slider("Hệ số công suất cosφ", 0.70, 1.00, 0.90, 0.01)
reserve_factor = st.sidebar.number_input("Hệ số dự phòng chiều dài dây", min_value=1.00, max_value=1.30, value=1.05, step=0.01)

st.sidebar.divider()
st.sidebar.subheader("Dây dẫn")
circuits = st.sidebar.selectbox("Số mạch thiết kế", [1, 2, 3, 4], index=1)
bundle = st.sidebar.number_input("Số dây/bó/pha thiết kế", min_value=1, max_value=6, value=int(defaults["bundle"]), step=1)
jkt = st.sidebar.number_input("Jkt tham khảo (A/mm2)", min_value=0.1, value=float(defaults["jkt"]), step=0.05)

st.sidebar.divider()
st.sidebar.subheader("Đường dây hiện trạng")
existing_voltage = st.sidebar.selectbox("Cấp điện áp hiện trạng", ["110kV", "220kV", "500kV"], index=["110kV", "220kV", "500kV"].index(voltage_level))
existing_conductor_code = st.sidebar.selectbox("Dây hiện trạng", CONDUCTOR_LIB["Mã dây"].tolist(), index=4)
existing_bundle = st.sidebar.number_input("Số dây/bó/pha hiện trạng", min_value=1, max_value=6, value=int(defaults["bundle"]), step=1)
existing_circuits = st.sidebar.selectbox("Số mạch hiện trạng", [1, 2, 3, 4], index=1)
existing_span = st.sidebar.number_input("Khoảng cột TB hiện trạng / tham khảo (m)", min_value=50, value=int(defaults["span_ref"]), step=10)

st.sidebar.divider()
st.sidebar.subheader("Tính số cột")
tower_count_mode = st.sidebar.radio("Chế độ tính số cột", ["Theo khoảng cột trung bình", "Nhập tổng số cột thực tế"], index=0)
manual_total_towers = st.sidebar.number_input("Tổng số cột thực tế", min_value=0, value=0, step=1, disabled=(tower_count_mode == "Theo khoảng cột trung bình"))

is_transit = connection_type.startswith("Transit")
boq_circuits = existing_circuits if is_transit else circuits
boq_bundle = existing_bundle if is_transit else bundle
boq_conductor = conductor_from_lib(existing_conductor_code) if is_transit else None

i_load = current_3phase(p_mw, u_kv, cos_phi)
s_econ = i_load / jkt if jkt > 0 else 0
i_per_wire_design = i_load / max(circuits * bundle, 1)
s_per_wire_design = s_econ / max(circuits * bundle, 1)
conductor_check = select_conductor(i_per_wire_design, s_per_wire_design)
recommended_df = conductor_check[conductor_check["Đạt sơ bộ"]]
recommended = recommended_df.iloc[0] if not recommended_df.empty else conductor_check.iloc[-1]
if boq_conductor is None:
    boq_conductor = recommended

i_per_wire_boq = i_load / max(boq_circuits * boq_bundle, 1)
s_per_wire_boq = s_econ / max(boq_circuits * boq_bundle, 1)
dv_percent = voltage_drop_percent(i_load, u_kv, length_km, float(boq_conductor["R ohm/km"]), float(boq_conductor["X ohm/km"]), cos_phi, boq_circuits, boq_bundle)
spans, calculated_tower_total = estimate_towers(length_km, existing_span)
tower_total = int(manual_total_towers) if tower_count_mode == "Nhập tổng số cột thực tế" and manual_total_towers > 0 else calculated_tower_total
conductor_length_km = length_km * boq_circuits * 3 * boq_bundle * reserve_factor
conductor_weight_t = conductor_length_km * float(boq_conductor["Khối lượng kg/km"]) / 1000
opgw_length_km = length_km * reserve_factor

st.title("⚡ BOQ sơ bộ đường dây 110kV–500kV")
st.caption("Bản v3: có logic transit, nhập số cột thực tế, phân bổ cột không lệch tổng, xuất Excel đúng BOQ tổng hợp.")
if is_transit:
    st.info("Transit/cắt xen: app dùng **dây hiện trạng** cho BOQ. Phần chọn dây tự động chỉ dùng để kiểm/cảnh báo, không tự thay size dây.")

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Dòng tải tổng", f"{i_load:,.0f} A")
c2.metric("Dòng trên 1 dây BOQ", f"{i_per_wire_boq:,.0f} A")
c3.metric("Dây BOQ", str(boq_conductor["Mã dây"]))
c4.metric("Số cột dùng tính", f"{tower_total:,}")
c5.metric("Dây dẫn", f"{conductor_weight_t:,.1f} tấn")

tab0, tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["00. Transit & Input", "01. Dây dẫn", "02. Cột", "03. Móng", "04. BOQ tổng hợp", "05. Check/RFI", "06. Hướng dẫn"])

with tab0:
    st.subheader("00. Transit & Input summary")
    input_summary = pd.DataFrame([
        {"Nhóm": "Dự án", "Thông số": "Tên dự án", "Giá trị": project_name, "Ghi chú": ""},
        {"Nhóm": "Dự án", "Thông số": "Cấp điện áp", "Giá trị": voltage_level, "Ghi chú": ""},
        {"Nhóm": "Dự án", "Thông số": "Kiểu đấu nối", "Giá trị": connection_type, "Ghi chú": "Transit: dây BOQ theo hiện trạng"},
        {"Nhóm": "Dự án", "Thông số": "Công suất truyền tải", "Giá trị": f"{p_mw} MW", "Ghi chú": ""},
        {"Nhóm": "Dự án", "Thông số": "Chiều dài", "Giá trị": f"{length_km} km", "Ghi chú": ""},
        {"Nhóm": "Hiện trạng", "Thông số": "Dây hiện trạng", "Giá trị": existing_conductor_code, "Ghi chú": "Dùng cho BOQ nếu transit"},
        {"Nhóm": "Hiện trạng", "Thông số": "Số mạch hiện trạng", "Giá trị": existing_circuits, "Ghi chú": ""},
        {"Nhóm": "Hiện trạng", "Thông số": "Số dây/bó/pha hiện trạng", "Giá trị": existing_bundle, "Ghi chú": ""},
        {"Nhóm": "Cột", "Thông số": "Chế độ tính số cột", "Giá trị": tower_count_mode, "Ghi chú": ""},
        {"Nhóm": "Cột", "Thông số": "Khoảng cột dùng tính", "Giá trị": f"{existing_span} m", "Ghi chú": "Nếu có tower schedule nên nhập tổng số cột thực tế"},
        {"Nhóm": "Cột", "Thông số": "Số cột dùng tính", "Giá trị": tower_total, "Ghi chú": ""},
    ])
    st.dataframe(input_summary, use_container_width=True, hide_index=True)
    transit_check = pd.DataFrame([
        {"Điều kiện": "Dòng trên 1 dây <= I cho phép dây BOQ", "Giá trị": f"{i_per_wire_boq:,.0f} <= {boq_conductor['I cho phép A']:,.0f} A", "Trạng thái": status(i_per_wire_boq <= float(boq_conductor["I cho phép A"])), "Ghi chú": "Nếu FAIL cần RFI/kiểm lại khả năng tải đường dây hiện trạng"},
        {"Điều kiện": "Tiết diện dây BOQ >= tiết diện kinh tế sơ bộ", "Giá trị": f"{boq_conductor['Tiết diện mm2']:,.0f} >= {s_per_wire_boq:,.0f} mm2", "Trạng thái": status(float(boq_conductor["Tiết diện mm2"]) >= s_per_wire_boq), "Ghi chú": "Transit có thể không đạt Jkt nhưng vẫn theo hiện trạng nếu được chấp thuận"},
        {"Điều kiện": "Tổn thất điện áp sơ bộ <= 5%", "Giá trị": f"{dv_percent:,.2f}%", "Trạng thái": status(dv_percent <= 5.0), "Ghi chú": "Ngưỡng 5% chỉ là check sơ bộ"},
    ])
    st.dataframe(transit_check, use_container_width=True, hide_index=True)

with tab1:
    st.subheader("01. Kiểm dây dẫn")
    st.dataframe(conductor_check, use_container_width=True, hide_index=True)
    st.write("**Dây khuyến nghị theo tính sơ bộ:**", recommended["Mã dây"])
    st.write("**Dây dùng BOQ:**", boq_conductor["Mã dây"])

with tab2:
    st.subheader("02. Nhập cơ cấu cột và tính khối lượng thép")
    st.warning("Nếu đã có tower schedule, nhập số lượng thực tế vào cột `Số lượng nhập tay`. Khi nhập tay toàn bộ, cột `Tỷ lệ %` không còn ý nghĩa.")
    default_mix = pd.DataFrame([
        {"Loại cột": "Đỡ thẳng", "Tỷ lệ %": 70, "Số lượng nhập tay": 0, "Khối lượng thép/cột trước mạ (tấn)": defaults["tower_weight"] * 0.75, "Hệ số mạ/hao hụt": 1.04, "Nguồn số liệu": "Assumption"},
        {"Loại cột": "Đỡ góc", "Tỷ lệ %": 8, "Số lượng nhập tay": 0, "Khối lượng thép/cột trước mạ (tấn)": defaults["tower_weight"] * 0.95, "Hệ số mạ/hao hụt": 1.04, "Nguồn số liệu": "Assumption"},
        {"Loại cột": "Néo góc", "Tỷ lệ %": 12, "Số lượng nhập tay": 0, "Khối lượng thép/cột trước mạ (tấn)": defaults["tower_weight"] * 1.35, "Hệ số mạ/hao hụt": 1.04, "Nguồn số liệu": "Assumption"},
        {"Loại cột": "Cột cuối", "Tỷ lệ %": 0, "Số lượng nhập tay": 2 if tower_total >= 2 else 0, "Khối lượng thép/cột trước mạ (tấn)": defaults["tower_weight"] * 1.50, "Hệ số mạ/hao hụt": 1.04, "Nguồn số liệu": "Assumption"},
        {"Loại cột": "Cột vượt", "Tỷ lệ %": 0, "Số lượng nhập tay": 0, "Khối lượng thép/cột trước mạ (tấn)": defaults["tower_weight"] * 2.50, "Hệ số mạ/hao hụt": 1.04, "Nguồn số liệu": "Assumption"},
        {"Loại cột": "Cột đặc biệt", "Tỷ lệ %": 0, "Số lượng nhập tay": 0, "Khối lượng thép/cột trước mạ (tấn)": defaults["tower_weight"] * 2.00, "Hệ số mạ/hao hụt": 1.04, "Nguồn số liệu": "Assumption"},
    ])
    tower_input = st.data_editor(default_mix, use_container_width=True, hide_index=True, num_rows="fixed", key="tower_input_v3")
    manual_sum = int(tower_input["Số lượng nhập tay"].sum())
    auto_total = max(0, tower_total - manual_sum)
    alloc = allocate_by_percent(auto_total, tower_input)
    rows = []
    for idx, row in tower_input.iterrows():
        manual_qty = int(row["Số lượng nhập tay"])
        qty = manual_qty if manual_qty > 0 else alloc.get(idx, 0)
        steel_before = qty * float(row["Khối lượng thép/cột trước mạ (tấn)"])
        steel_after = steel_before * float(row["Hệ số mạ/hao hụt"])
        rows.append({"Loại cột": row["Loại cột"], "Số lượng": qty, "Cách tính": "Nhập tay" if manual_qty > 0 else "Phân bổ tỷ lệ", "KL thép/cột trước mạ (tấn)": float(row["Khối lượng thép/cột trước mạ (tấn)"]), "Tổng thép trước mạ (tấn)": steel_before, "Hệ số mạ/hao hụt": float(row["Hệ số mạ/hao hụt"]), "Tổng thép sau mạ/hao hụt (tấn)": steel_after, "Nguồn số liệu": row.get("Nguồn số liệu", "Assumption")})
    tower_calc = pd.DataFrame(rows)
    if manual_sum > tower_total:
        st.error(f"Số lượng nhập tay ({manual_sum}) đang lớn hơn tổng số cột dùng tính ({tower_total}). Cần sửa lại.")
    elif tower_calc["Số lượng"].sum() != tower_total:
        st.error("Tổng số lượng cột sau phân bổ chưa khớp tổng số cột dùng tính. Cần kiểm tra lại tỷ lệ %.")
    else:
        st.success(f"Tổng số cột sau phân bổ = {tower_calc['Số lượng'].sum()} cột, khớp với số cột dùng tính.")
    st.dataframe(tower_calc, use_container_width=True, hide_index=True)

with tab3:
    st.subheader("03. Móng, tiếp địa")
    foundation_default = pd.DataFrame([
        {"Hạng mục": "Đào đất đá hố móng", "Đơn vị": "m3/cột", "Suất KL": defaults["foundation_exc"], "Nguồn số liệu": "Assumption"},
        {"Hạng mục": "Lấp đất hố móng", "Đơn vị": "m3/cột", "Suất KL": defaults["foundation_exc"] * 0.75, "Nguồn số liệu": "Assumption"},
        {"Hạng mục": "Bê tông lót M50/M100", "Đơn vị": "m3/cột", "Suất KL": defaults["foundation_conc"] * 0.08, "Nguồn số liệu": "Assumption"},
        {"Hạng mục": "Bê tông móng M150/M200", "Đơn vị": "m3/cột", "Suất KL": defaults["foundation_conc"], "Nguồn số liệu": "Assumption"},
        {"Hạng mục": "Cốt thép móng", "Đơn vị": "tấn/cột", "Suất KL": defaults["rebar"], "Nguồn số liệu": "Assumption"},
        {"Hạng mục": "Bu lông neo", "Đơn vị": "tấn/cột", "Suất KL": defaults["anchor"], "Nguồn số liệu": "Assumption"},
        {"Hạng mục": "Tiếp địa", "Đơn vị": "tấn/cột", "Suất KL": defaults["grounding"], "Nguồn số liệu": "Assumption"},
        {"Hạng mục": "San gạt/kè/rãnh", "Đơn vị": "m3/cột", "Suất KL": 0.0, "Nguồn số liệu": "Nếu có"},
    ])
    foundation_input = st.data_editor(foundation_default, use_container_width=True, hide_index=True, num_rows="fixed", key="foundation_input_v3")
    foundation_input["Số cột áp dụng"] = tower_calc["Số lượng"].sum()
    foundation_input["Tổng khối lượng"] = foundation_input["Suất KL"] * foundation_input["Số cột áp dụng"]
    st.dataframe(foundation_input, use_container_width=True, hide_index=True)

# Build BOQ/state before showing/exporting
boq_rows = [
    {"Nhóm": "Dây dẫn", "Hạng mục": f"Dây dẫn {boq_conductor['Mã dây']}", "Đơn vị": "km", "Khối lượng": conductor_length_km, "Nguồn số liệu": "Transit: dây hiện trạng / Đấu nối mới: dây chọn sơ bộ", "Ghi chú": f"{boq_circuits} mạch x 3 pha x {boq_bundle} dây/bó x hệ số {reserve_factor}"},
    {"Nhóm": "Dây dẫn", "Hạng mục": f"Khối lượng dây dẫn {boq_conductor['Mã dây']}", "Đơn vị": "tấn", "Khối lượng": conductor_weight_t, "Nguồn số liệu": "Chiều dài dây x kg/km", "Ghi chú": "kg/km lấy từ thư viện, cần thay bằng catalog"},
    {"Nhóm": "Dây chống sét/OPGW", "Hạng mục": "OPGW / dây chống sét", "Đơn vị": "km", "Khối lượng": opgw_length_km, "Nguồn số liệu": "Chiều dài tuyến x dự phòng", "Ghi chú": "Cần xác nhận số sợi OPGW/dây chống sét"},
    {"Nhóm": "Cột", "Hạng mục": "Tổng số cột", "Đơn vị": "cột", "Khối lượng": tower_calc["Số lượng"].sum(), "Nguồn số liệu": tower_count_mode, "Ghi chú": f"Khoảng cột dùng tính: {existing_span} m"},
    {"Nhóm": "Cột", "Hạng mục": "Thép cột trước mạ", "Đơn vị": "tấn", "Khối lượng": tower_calc["Tổng thép trước mạ (tấn)"].sum(), "Nguồn số liệu": "Bảng cơ cấu cột", "Ghi chú": ""},
    {"Nhóm": "Cột", "Hạng mục": "Thép cột sau mạ/hao hụt", "Đơn vị": "tấn", "Khối lượng": tower_calc["Tổng thép sau mạ/hao hụt (tấn)"].sum(), "Nguồn số liệu": "Bảng cơ cấu cột", "Ghi chú": ""},
]
for _, row in foundation_input.iterrows():
    boq_rows.append({"Nhóm": "Móng & xây lắp" if row["Hạng mục"] != "Tiếp địa" else "Tiếp địa", "Hạng mục": row["Hạng mục"], "Đơn vị": row["Đơn vị"].split("/")[0], "Khối lượng": row["Tổng khối lượng"], "Nguồn số liệu": row["Nguồn số liệu"], "Ghi chú": f"Suất KL {row['Suất KL']} {row['Đơn vị']} x {row['Số cột áp dụng']} cột"})
boq = normalize_boq(pd.DataFrame(boq_rows))
boq_summary = boq.groupby(["Nhóm", "Đơn vị"], as_index=False)["Khối lượng"].sum()

check_df = pd.DataFrame([
    {"Nhóm": "Transit", "Điều kiện": "Nếu transit thì dây BOQ phải theo hiện trạng", "Trạng thái": "PASS" if is_transit else "N/A", "RFI/Ghi chú": "Cần hồ sơ hiện trạng xác nhận dây dẫn, số mạch, số dây/bó, phụ kiện"},
    {"Nhóm": "Dây dẫn", "Điều kiện": "Dòng trên 1 dây <= I cho phép", "Trạng thái": status(i_per_wire_boq <= float(boq_conductor["I cho phép A"])), "RFI/Ghi chú": "Nếu FAIL cần kiểm lại khả năng tải hoặc phương án đấu nối"},
    {"Nhóm": "Dây dẫn", "Điều kiện": "Điện áp rơi sơ bộ <= 5%", "Trạng thái": status(dv_percent <= 5.0), "RFI/Ghi chú": "Ngưỡng sơ bộ; cần thay bằng tiêu chí dự án"},
    {"Nhóm": "Cột", "Điều kiện": "Tổng cột phân bổ khớp tổng cột dùng tính", "Trạng thái": status(tower_calc["Số lượng"].sum() == tower_total), "RFI/Ghi chú": "Nếu FAIL cần sửa số lượng nhập tay/tỷ lệ"},
    {"Nhóm": "Cột", "Điều kiện": "Khối lượng thép/cột có nguồn chính thức", "Trạng thái": "OPEN-RFI" if tower_input["Nguồn số liệu"].astype(str).str.contains("Assumption", case=False).any() else "PASS", "RFI/Ghi chú": "Cần bảng trọng lượng cột/bản vẽ chế tạo"},
    {"Nhóm": "Móng", "Điều kiện": "Suất móng có nguồn chính thức", "Trạng thái": "OPEN-RFI" if foundation_input["Nguồn số liệu"].astype(str).str.contains("Assumption", case=False).any() else "PASS", "RFI/Ghi chú": "Cần bản vẽ móng/phụ lục tính móng"},
])

with tab4:
    st.subheader("04. BOQ tổng hợp sơ bộ")
    st.dataframe(boq, use_container_width=True, hide_index=True)
    st.write("**Tổng hợp theo nhóm/đơn vị**")
    st.dataframe(boq_summary, use_container_width=True, hide_index=True)
    export_sheets = {"00_Input": input_summary, "01_Transit_Check": transit_check, "02_Day_Dan_Check": conductor_check, "03_Co_Cau_Cot_Input": tower_input, "04_Cot_Calc": tower_calc, "05_Mong_Tiep_Dia": foundation_input, "06_BOQ_Tong_Hop": boq, "07_BOQ_Summary": boq_summary, "08_Check_RFI": check_df, "09_Thu_Vien_Day": CONDUCTOR_LIB}
    excel_bytes = build_excel(export_sheets)
    st.download_button("⬇️ Xuất Excel đúng theo BOQ tổng hợp đang hiển thị", data=excel_bytes, file_name=f"BOQ_Duong_day_{voltage_level}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tab5:
    st.subheader("05. Check / RFI")
    st.dataframe(check_df, use_container_width=True, hide_index=True)

with tab6:
    st.subheader("06. Hướng dẫn nhập liệu")
    guide = pd.DataFrame([
        {"Mục": "Kiểu đấu nối", "Nhập ở đâu": "Sidebar", "Lấy số liệu ở đâu": "Thỏa thuận đấu nối / phương án đấu nối", "Lưu ý": "Transit thì dây BOQ theo hiện trạng"},
        {"Mục": "Dây hiện trạng", "Nhập ở đâu": "Sidebar", "Lấy số liệu ở đâu": "Hồ sơ hoàn công / khảo sát / EVN", "Lưu ý": "Ví dụ ACSR 2x330 thì chọn ACSR-330 và số dây/bó = 2"},
        {"Mục": "Khoảng cột", "Nhập ở đâu": "Sidebar", "Lấy số liệu ở đâu": "Mặt cắt dọc / tower schedule / hiện trạng", "Lưu ý": "Nếu có số cột thực tế thì chọn chế độ nhập tổng số cột"},
        {"Mục": "Cơ cấu cột", "Nhập ở đâu": "Tab 02", "Lấy số liệu ở đâu": "Tower schedule", "Lưu ý": "Nhập số lượng tay để tránh tỷ lệ giả định"},
        {"Mục": "Khối lượng thép/cột", "Nhập ở đâu": "Tab 02", "Lấy số liệu ở đâu": "Bảng trọng lượng cột / bản vẽ chế tạo", "Lưu ý": "Không lấy mặc định để chốt BOQ"},
        {"Mục": "Móng", "Nhập ở đâu": "Tab 03", "Lấy số liệu ở đâu": "Bản vẽ móng / phụ lục tính móng", "Lưu ý": "Tách bê tông lót, bê tông móng, cốt thép, bu lông neo"},
        {"Mục": "Xuất Excel", "Nhập ở đâu": "Tab 04", "Lấy số liệu ở đâu": "Nút bên dưới", "Lưu ý": "Bản v3 xuất đúng theo BOQ đang hiển thị"},
    ])
    st.dataframe(guide, use_container_width=True, hide_index=True)
